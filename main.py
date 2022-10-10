# This is a sample Python script.
import sys

# Press Shift+F10 to execute it or replace it with your code.

# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


import pandas as pd

START_SKIP_ROWS = 4

SECTIONS_NUM = 16
NO_OF_LOOPS = 11
NO_OF_DEVICES = 8
def read_excel():
    df = pd.read_excel(r'd:\excel\DAILY REPORT.xlsx', sheet_name='DUT1', skiprows=4)

    df = pd.DataFrame(df, columns=['Section', 'Total Iterations', 'Total Failed'])

    return df


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.

    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.

if __name__ == '__main__':
    from openpyxl import load_workbook

    #xls = xlrd.open_workbook(r'd:\excel\DAILY REPORT.xlsx', on_demand=True)
    wb = load_workbook(r'd:\excel\DAILY REPORT.xlsx', read_only=True, keep_links=False)
    # print the sheet names contains "DUT" in the name
    devices = len([sheet for sheet in wb.sheetnames if "DUT" in sheet])
    print(devices)

    #print(wb.sheetnames)  # <- remeber: xlrd sheet_names is a function, not a property
    sys.exit(1)
    print_hi('PyCharm')

    #x = read_excel()

    df = pd.read_excel(r'd:\excel\DAILY REPORT.xlsx', sheet_name='DUT1', skiprows=4)

    # df = pd.DataFrame(df, columns=['Section', 'Total Iterations', 'Total Failed'])

    #print(df.keys())

    #loop = df.iloc[0:16]
    #print("loop:" + str(loop))
    #print(type(loop))
    #print(loop[['Section', 'Total Iterations', 'Total Failed']])
    #loop = loop['Section', 'Total Iterations', 'Total Failed']
    #print(loop)

    #print(loop['Total Failed'].sum())
    #loop = df.iloc[19:35]
    #print(loop[['Section', 'Total Iterations', 'Total Failed']])

    #loop = df.iloc[38:54]
    #print(loop[['Section', 'Total Iterations', 'Total Failed']])

    all_dev_list = []
    for dev in range(1, NO_OF_DEVICES+1):
        df = pd.read_excel(r'd:\excel\DAILY REPORT.xlsx', sheet_name='DUT%s'%dev, skiprows=4)
        b=16
        loop_list = []
        for i in range(0, NO_OF_LOOPS):
            loop = df.iloc[19 * i: b ]
            b = b + 19
            #print(loop[['Section', 'Total Iterations', 'Total Failed']])
            d = {}
            #d['Section'] = loop['Section'].values
            #d['Total Iterations'] = loop['Total Iterations'].values
            #d['Total Failed'] = loop['Total Failed'].values
            d['data'] = loop[['Section', 'Total Iterations', 'Total Failed']]
            d['loop_sum failed'] = loop['Total Failed'].sum()
            #d['device_no'] = 1
            loop_list.append(d)
        dict = {}
        dict['device_no'] = dev
        dict['loops'] = loop_list
        dict["device_sum_failed"] = sum([x['loop_sum failed'] for x in loop_list])
        all_dev_list.append(dict)
    print("*"*100)
    #print(all_dev_list)
    for dev in all_dev_list:
        print(f'\nDevice Name:, {dev["device_no"]}  Failed: {dev["device_sum_failed"]}')