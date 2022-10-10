import argparse
import pandas as pd
import sys
pd.options.mode.chained_assignment = None  # default='warn'

#Constants
BEST_X_DEVICES = 5
DAILY_REPORT_FILE = r'd:\excel\DAILY REPORT.xlsx'
DEVICE_RECOGNIZER_KEYWORD = "DUT"
DEV_SHEET_NAME = '%s{}'%DEVICE_RECOGNIZER_KEYWORD
DEV_MAPPER_SHEET_NAME = 'issues'
START_SKIP_ROWS = 4
SECTIONS_NUM = 16
NO_OF_LOOPS = 11
NO_OF_DEVICES = 8
ROWS_BETWEEN_LOOPS = 3
NO_ROWS_TO_READ = ROWS_BETWEEN_LOOPS + SECTIONS_NUM

# Constants for Final Report
FINAL_ROWS_BETWEEN_DEVICES = 8
FINAL_START_SKIP_ROWS = 19
FINAL_REPORT_FILE = r'd:\excel\Final Report.xlsx'
FINAL_REPORT_SHEET_NAME = '400Hr. Completion Percentage'

class DailyReportParser:
    def __init__(self, daily_report_path, final_report_path, best_x_devices):
        self.daily_report_path = daily_report_path
        self.best_x_devices = best_x_devices
        self.final_report_path = final_report_path

    def collect_devices_info(self, verbose=True):
        """
        Collects the devices info from the daily report file.

        :param verbose: if True, prints the devices info.
        :return: A list of the devices info.
        """
        all_dev_list = []
        for dev in range(1, self.get_devices_number() + 1):
            df = pd.read_excel(self.daily_report_path, sheet_name=DEV_SHEET_NAME.format(dev), skiprows=START_SKIP_ROWS, index_col=0)
            b = SECTIONS_NUM
            loop_list = []
            for i in range(0, NO_OF_LOOPS):
                loop = df.iloc[NO_ROWS_TO_READ * i: b]
                b = b + NO_ROWS_TO_READ
                loop["Total Success"] = loop["Total Iterations"] - loop["Total Failed"]
                d = {}
                d['data'] = loop[['Section', 'Total Iterations', 'Total Failed', 'Total Success']]
                d['loop_sum failed'] = loop['Total Failed'].sum()

                loop_list.append(d)
            dev_info = {}
            dev_info['device_no'] = dev
            dev_info['loops'] = loop_list
            dev_info["device_sum_failed"] = sum([x['loop_sum failed'] for x in loop_list])
            all_dev_list.append(dev_info)
        if verbose:
            for dev in all_dev_list:
                print(f'\nDevice Name:{dev["device_no"]}  Total Failed: {dev["device_sum_failed"]}')
        return  all_dev_list

    def get_best_x_devices(self, dev_list, x=None, verbose=True):
        """
        Returns the best x devices from the list of devices.
        best x devices are the devices with the lowest number of failed tests.
        :param dev_list: the list of collected devices.
        :param x: the number of best devices to return.
        :param verbose: if True, prints the best x devices.
        :return: A list of the best x devices.
        """
        x = x if x else self.best_x_devices
        dev_list.sort(key=lambda x: x['device_sum_failed'])
        if verbose:
            print("*" * 100)
            print(f"Best {x} devices with the lowest failed:")
            for dev in dev_list[:x]:
                print(f'\nDevice Name:{dev["device_no"]}  Total Failed: {dev["device_sum_failed"]}')
            print("*" * 100)
        return dev_list[:x]

    def get_devices_number(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.daily_report_path, read_only=True, keep_links=False)
        return len([sheet for sheet in wb.sheetnames if DEVICE_RECOGNIZER_KEYWORD in sheet])

    def write_device_mapper(self, dev_list):
        """
        Writes the best x devices to the device mapper sheet.
        :param dev_list: the list of the best x devices.
        :return: None. It writes to the excel file directly.
        """
        print(f"* Writing the device mapper table into the sheet '{DEV_MAPPER_SHEET_NAME}' ...")
        data = {'Final Report': ['Device#{}'.format(i) for i in range(1, len(dev_list) + 1)],
                'Daily Report': [DEV_SHEET_NAME.format(x['device_no']) for x in dev_list]}
        df = pd.DataFrame(data)
        with pd.ExcelWriter(self.daily_report_path,
                            mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=DEV_MAPPER_SHEET_NAME, index=False, startrow=25)


    def write_into_final(self, best_dev_list):
        """
        Writes the best x devices data to the final report file.
        :param best_dev_list: the list of the best x devices.
        :return: None. It writes to the excel file directly.
        """
        print(f"* Writing the best {self.best_x_devices} devices into the final report '{self.final_report_path}' sheet '{FINAL_REPORT_SHEET_NAME}' ...")
        start_row = FINAL_START_SKIP_ROWS
        for dev in best_dev_list:
            df = pd.DataFrame()
            df["Total Iterations"] = dev['loops'][0]['data']['Total Iterations']
            for loop in range(0, len(dev['loops'])):
                df["loop%s" % loop] = dev["loops"][loop]["data"][['Total Success']]
            with pd.ExcelWriter(self.final_report_path,
                                mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=FINAL_REPORT_SHEET_NAME, index=False, startcol=2, startrow=start_row, header=False)
            start_row = start_row + SECTIONS_NUM + FINAL_ROWS_BETWEEN_DEVICES ## the number of rows to be skipped between devices tables



def parse_args():
    parser = argparse.ArgumentParser()

    parser.add_argument("-d", "--daily_path", action="store",
                        dest="daily_report_path", default=DAILY_REPORT_FILE,
                        help="Daily report excel file path")
    parser.add_argument("-f", "--final_path", action="store",
                        dest="final_report_path", default=FINAL_REPORT_FILE,
                        help="Final report excel file path")
    parser.add_argument("-b", "--best_devices", action="store",
                        dest="best_x_devices", default=BEST_X_DEVICES,
                        help="Number of needed best devices", type=int)

    args = parser.parse_args(sys.argv[1:])
    return DailyReportParser(args.daily_report_path, args.final_report_path, args.best_x_devices)

if __name__ == '__main__':
    try:
        script = parse_args()
        print("* Starting the script...")
        print(f"* Reading the excel file '{script.daily_report_path}' and collect the best {script.best_x_devices} devices...")
        dev_info = script.collect_devices_info(verbose=False)
        x= script.get_best_x_devices(dev_info, verbose=False)
        script.write_device_mapper(x)
        script.write_into_final(x)
        print("Done!")
    except Exception as e:
        print(e)
        print("Error: Something went wrong. Please check the input files and try again.")
        sys.exit(1)





